import openpyxl
import ipaddress
import socket
import os

wb = openpyxl.load_workbook(input('please enter your excel name:\n'))
sheet = wb.active
column : int = int(input('which column: (example : 1)\n'))
for row in range(1, sheet.max_row +1):
    ip = sheet.cell(row=row, column=column).value
    print('================================')
    print('check ip: ' + ip + '\n')
    try:
        host_ip = socket.gethostbyname(ip) 
        ipaddress.ip_address(host_ip)
        cmd_ping = os.system('ping -c 3 '+ host_ip)
        if cmd_ping == 0:
            print('\n' + 'host is up\n')
        else:
            print('\n' + 'host is down\n')
    except:
        print(f"input is: {ip} \nInvalid IP address")
        
